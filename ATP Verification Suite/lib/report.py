"""Write execution_report.xlsx and summary.json."""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

_REPO = Path(__file__).resolve().parents[2]
if str(_REPO) not in sys.path:
    sys.path.insert(0, str(_REPO))

from utils.project_identity import PROJECT_DISPLAY_NAME  # noqa: E402


HEADERS = [
    "Test ID",
    "Module",
    "Description",
    "Expected Result",
    "Actual Result",
    "Status",
    "Failure Reason",
    "Screenshot",
    "Execution Time",
]


def write_reports(repo_root: Path, summary: dict) -> tuple[Path, Path]:
    reports = repo_root / "reports"
    reports.mkdir(parents=True, exist_ok=True)
    xlsx = reports / "execution_report.xlsx"
    js = reports / "summary.json"

    wb = Workbook()
    ws = wb.active
    ws.title = "Execution"
    ws["A1"] = f"{PROJECT_DISPLAY_NAME} - ATP Verification Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="333333")
    header_row = 2
    for col, _ in enumerate(HEADERS, 1):
        cell = ws.cell(header_row, col)
        cell.font = header_font
        cell.fill = header_fill

    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")

    for row in summary.get("rows", []):
        values = [
            row.get("test_id", ""),
            row.get("module", ""),
            row.get("verification") or row.get("description", ""),
            row.get("expected", ""),
            row.get("actual", ""),
            row.get("status", ""),
            row.get("failure_reason", ""),
            row.get("screenshot", ""),
            row.get("execution_time", ""),
        ]
        ws.append(values)
        status_cell = ws.cell(ws.max_row, 6)
        if row.get("status") == "PASS":
            status_cell.fill = pass_fill
        elif row.get("status") == "FAIL":
            status_cell.fill = fail_fill

    # Summary sheet
    sm = wb.create_sheet("Summary")
    sm["A1"] = f"{PROJECT_DISPLAY_NAME} - Summary"
    sm["A1"].font = Font(bold=True, size=14)
    sm.append(["Metric", "Value"])
    for k in (
        "total_verifications",
        "passed",
        "failed",
        "pass_percent",
        "execution_time_sec",
    ):
        sm.append([k, summary.get(k)])
    sm.append(["failed_modules", ", ".join(summary.get("failed_modules", []))])
    sm.append(["project", PROJECT_DISPLAY_NAME])

    wb.save(xlsx)
    summary_out = dict(summary)
    summary_out["project"] = PROJECT_DISPLAY_NAME
    js.write_text(json.dumps(summary_out, indent=2, ensure_ascii=False), encoding="utf-8")
    return xlsx, js
