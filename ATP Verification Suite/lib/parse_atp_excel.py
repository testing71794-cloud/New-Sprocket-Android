"""Parse HP Sprocket ATP Excel into normalized stage catalogs."""
from __future__ import annotations

import json
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook

SUITE = Path(__file__).resolve().parents[1]
CATALOG = SUITE / "catalog"
DEFAULT_XLSX = CATALOG / "Hp_new_sprocket_ATP_Sheet.xlsx"

# Fine-grained ATP "Module" cells → coarse stage used for one-flow-per-module runs.
# Order matters: more specific product areas before generic keywords (e.g. collage before gallery).
STAGE_RULES: list[tuple[str, str]] = [
    (r"splash screen|^splash\b", "Splash"),
    (r"onboarding|sign up|log[\s_-]?in|auth|forgot password|privacy policy|app launch & auth|create account", "Onboarding"),
    (r"home screen|application launch and home|notification permission", "Home"),
    (r"collage|drag & drop|layout selection|image reorder|image selection|duplicate selection|remove selection|folder navigation", "Collage"),
    (r"precut|pre-cut|warning pop-up|exit & save", "PreCut"),
    (r"quick print|facebook|select mode|select gallery|sort menu|no photos|\bgallery\b", "QuickPrint"),
    (r"video frame|\bvideos\b|cr_05", "Video"),
    (r"print preview|print action|print flow|print loader|print complete|print tip|print queue|print interruption|print state|print ui|saved to gallery|multicopy|deleting from queue|single photo print|multiple photo|multiple copy|blocker|resume queue|active printing", "Printing"),
    (r"firmware", "Firmware"),
    (r"\btile\b|\btiles\b", "TilePrint"),
    (r"photobooth|countdown|capture view|flash logic|\bcamera\b|\btimer\b|photo id|aspect ratio", "Camera"),
    (r"sprocket ai|ai tool|text-to-image|replace object|restore photo|extend image|generate background|show original|prompting", "AI"),
    (r"pre connection|connection flow|discovery|bluetooth|wi-fi|already added|printer settings|manage printers|printers detail|printer 200|reconnection|\bprinter\b", "Connection"),
    (r"settings|account|logout|change password|personal|legal|privacy|version|permissions|display|\bdata\b", "Settings"),
    (r"memory error|transmit error|low voltage|print error|high temperature|paper|cover|battery|temperature|\bsystem\b|ad_\d|ios_\d|error handling", "Alerts"),
    (r"image editor|edit screen|\bcrop\b|gesture|snapping|photo fit|edit tip|save photo|video editing|custom sdk|\bfilter|\bsticker|\bframe\b|\bundo\b|\bredo\b", "Editor"),
    (r"navigation|ui layout|carousel|functionality|controls|hardware|automation|interaction|ui/|ui coverage|ui block|ui backdrop|animation|performance|guard|cross-module|toast|instructional|workflow|preview|initialization|logic|validation|cancellation|saving|selection|result|generation|paper type", "General"),
]


def _stage_for(module: str, section: str | None = None) -> str:
    """Map ATP Module → stage. Prefer Module; only fall back to section for non-splash cues."""
    mod = (module or "").strip().lower()
    for pattern, stage in STAGE_RULES:
        if re.search(pattern, mod, re.I):
            return stage
    # Section headers are noisy (e.g. "ONBOARDING SPLASH SCREEN") — use only as weak hint,
    # and never assign Splash from section alone.
    sec = (section or "").strip().lower()
    if sec:
        for pattern, stage in STAGE_RULES:
            if stage == "Splash":
                continue
            if re.search(pattern, sec, re.I):
                return stage
    return "General"


def _parse_main(ws) -> list[dict]:
    cases: list[dict] = []
    current = None
    section = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=9, values_only=True), 1):
        vals = list(row) + [None] * 9
        tid, module, desc, steps, pn, expected, *_ = vals[:9]
        if tid and not module and not steps:
            section = str(tid).strip()
            continue
        tid_s = str(tid or "").strip()
        mod_s = str(module or "").strip()
        if tid_s.lower() in {"test id", "tc id", "module"} or mod_s.lower() == "module":
            continue
        if tid and module:
            current = {
                "test_id": str(tid).strip(),
                "module": str(module).strip(),
                "section": section,
                "description": (str(desc).strip() if desc else ""),
                "steps": [str(steps).strip()] if steps else [],
                "type": (str(pn).strip() if pn else ""),
                "expected": (str(expected).strip() if expected else ""),
                "sheet": "New Sprocket Final ATP",
                "row": i,
            }
            cases.append(current)
            continue
        if current and steps and not tid:
            current["steps"].append(str(steps).strip())
            if expected and not current["expected"]:
                current["expected"] = str(expected).strip()
            if pn and not current["type"]:
                current["type"] = str(pn).strip()
    return cases


def _parse_simple(ws, sheet: str) -> list[dict]:
    cases: list[dict] = []
    headers = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=8, values_only=True), 1):
        vals = list(row)
        if not any(vals):
            continue
        if headers is None:
            joined = " ".join(str(v).lower() for v in vals if v)
            if "tc id" in joined or "test id" in joined:
                headers = True
            continue
        if not vals[0] or str(vals[0]).strip().upper() in {"TC ID", "TEST ID", "NONE"}:
            continue
        steps_raw = str(vals[3]) if vals[3] else ""
        cases.append(
            {
                "test_id": str(vals[0]).strip(),
                "module": str(vals[1]).strip() if vals[1] else sheet,
                "section": sheet,
                "description": str(vals[2]).strip() if vals[2] else "",
                "steps": [s.strip() for s in steps_raw.split("\n") if s and s.strip()],
                "type": str(vals[4]).strip() if vals[4] else "",
                "expected": str(vals[5]).strip() if vals[5] else "",
                "sheet": sheet,
                "row": i,
            }
        )
    return cases


def extract_visible_candidates(expected: str, description: str = "") -> list[str]:
    """Pull quoted UI labels and known keywords from expected/description text."""
    blob = f"{expected}\n{description}"
    found: list[str] = []

    def _looks_like_atp_id(t: str) -> bool:
        # Skip sheet labels like "ON_05F_Log In - Invalid Email..." that are not on-screen text
        if re.match(r"^(ON|SU|LI|SP|GA|CO|COL|QP|PM|SG|LO|SL|CR|AD)[_\s]", t, re.I):
            return True
        if re.search(r"\bTC_[A-Z0-9_]+\b", t):
            return True
        if "_" in t and re.search(r"[A-Z]{2,}_\d", t):
            return True
        return False

    for m in re.finditer(r"[\"'“”]([^\"'“”]{2,80})[\"'“”]", blob):
        t = m.group(1).strip()
        if not t or _looks_like_atp_id(t):
            continue
        if t not in found:
            found.append(t)
    # Common Sprocket labels if mentioned without quotes
    known = [
        "Welcome to the new Sprocket!",
        "Swipe to continue.",
        "Get Started!",
        "Sign up",
        "Sign Up Later",
        "I already have an account",
        "Log In",
        "Continue without Logging In",
        "Collage Maker",
        "Quick Print",
        "Select 2 to 4 photos",
        "Drag & Drop to Reorder",
        "Copies",
        "Add Printer",
        "No Printers Added",
        "Allow all",
        "Don't allow",
        "Create",
        "Printer",
        "More",
        "Next",
        "Skip",
        "Enable Bluetooth",
        "Add New Printer",
        "Searching for Printers",
        "Invalid Email",
        "Invalid email",
        "password",
    ]
    low = blob.lower()
    for k in known:
        if k.lower() in low and k not in found:
            found.append(k)
    return found[:12]


def _infer_kind(text: str, expected: str) -> str:
    """Map expected wording to the closest helper kind."""
    blob = f"{text} {expected}".lower()
    if any(k in blob for k in ("toast", "snackbar")):
        return "toast"
    if any(k in blob for k in ("dialog", "pop-up", "popup", "alert")):
        return "dialog"
    if any(k in blob for k in ("permission", "allow all", "don't allow", "while using")):
        return "permission"
    if any(k in blob for k in ("disabled", "greyed", "grayed", "not enabled", "inactive")):
        return "disabled"
    if any(k in blob for k in ("enabled", "tappable", "active button")):
        return "enabled"
    if any(k in blob for k in ("logo", "icon", "image", "illustration", "animation")):
        return "image"
    if any(k in blob for k in ("navigate", "navigates", "opens", "transitions", "lands on")):
        return "navigation"
    if any(k in blob for k in ("not visible", "should not", "disappears", "hidden")):
        return "not_visible"
    return "visible"


def build_verifications(case: dict) -> list[dict]:
    """Turn one ATP case into a list of non-fatal verification specs."""
    vers: list[dict] = []
    base = {
        "test_id": case["test_id"],
        "module": case["module"],
        "stage": case["stage"],
        "description": case["description"],
        "expected": case["expected"],
    }
    expected = case.get("expected", "")
    candidates = extract_visible_candidates(expected, case.get("description", ""))
    for text in candidates:
        kind = _infer_kind(text, expected)
        label = {
            "visible": "Verify visible",
            "not_visible": "Verify not visible",
            "enabled": "Verify enabled",
            "disabled": "Verify disabled",
            "image": "Verify image/logo",
            "navigation": "Verify navigation marker",
            "toast": "Verify toast",
            "dialog": "Verify dialog",
            "permission": "Verify permission UI",
        }.get(kind, "Verify")
        vers.append(
            {
                **base,
                "verification": f"{label}: {text}",
                "kind": kind,
                "selector": text,
            }
        )
    # Splash-specific known markers from expected/description wording
    splash_markers = [
        ("splash", "Welcome to the new Sprocket!"),
        ("animation", "Welcome to the new Sprocket!"),
        ("swipe to continue", "Swipe to continue."),
        ("get started", "Get Started!"),
    ]
    low_blob = f"{case.get('expected','')} {case.get('description','')}".lower()
    if case.get("stage") == "Splash" or "splash" in case.get("module", "").lower():
        for needle, label in splash_markers:
            if needle in low_blob and not any(v.get("selector") == label for v in vers):
                vers.append(
                    {
                        **base,
                        "verification": f"Verify visible: {label}",
                        "kind": "visible",
                        "selector": label,
                    }
                )
    if not vers:
        vers.append(
            {
                **base,
                "verification": "Manual/heuristic check — review expected result on current screen",
                "kind": "manual_note",
                "selector": None,
            }
        )
    return vers


def parse_workbook(xlsx: Path = DEFAULT_XLSX) -> dict:
    wb = load_workbook(xlsx, data_only=True)
    cases = _parse_main(wb["New Sprocket Final ATP"])
    if "Connection Flow" in wb.sheetnames:
        cases += _parse_simple(wb["Connection Flow"], "Connection Flow")
    if "Printers detail pages" in wb.sheetnames:
        cases += _parse_simple(wb["Printers detail pages"], "Printers detail pages")

    # Drop junk
    cases = [c for c in cases if c["test_id"] and c["module"].lower() != "module"]

    for c in cases:
        c["stage"] = _stage_for(c["module"], c.get("section"))
        c["verifications"] = build_verifications(c)

    stages: OrderedDict[str, list] = OrderedDict()
    for c in cases:
        stages.setdefault(c["stage"], []).append(c)

    return {
        "source": str(xlsx),
        "total_cases": len(cases),
        "stages": {k: len(v) for k, v in stages.items()},
        "cases": cases,
        "by_stage": {k: v for k, v in stages.items()},
    }


def main() -> None:
    CATALOG.mkdir(parents=True, exist_ok=True)
    data = parse_workbook()
    out = CATALOG / "atp_cases.json"
    # Slim file for runners (without huge by_stage duplication of objects — keep cases + stages index)
    slim = {
        "source": data["source"],
        "total_cases": data["total_cases"],
        "stages": data["stages"],
        "cases": data["cases"],
    }
    out.write_text(json.dumps(slim, indent=2, ensure_ascii=False), encoding="utf-8")
    # Per-stage catalogs
    stage_dir = CATALOG / "stages"
    stage_dir.mkdir(exist_ok=True)
    for stage, cases in data["by_stage"].items():
        (stage_dir / f"{stage}.json").write_text(
            json.dumps({"stage": stage, "cases": cases}, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    print(f"wrote {out} cases={data['total_cases']} stages={len(data['stages'])}")
    for s, n in data["stages"].items():
        print(f"  {s}: {n}")


if __name__ == "__main__":
    main()
