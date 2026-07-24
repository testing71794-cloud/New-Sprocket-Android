"""
Run all Maestro flows in one ATP TestCase Flows module; continue after failures.
Writes reports/module_runs/<module>_summary.json and appends to module_runs/index.json.

Multi-device: pass comma-separated serials to --device (or repeat --device).
Flows are pulled from a shared queue (dynamic parallel), one worker per phone.
"""
from __future__ import annotations

import argparse
import json
import os
import queue
import subprocess
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
ATP = REPO / "ATP TestCase Flows"
OUT = REPO / "reports" / "module_runs"
APP_ID = "com.hp.impulse.sprocket"


def discover_flows(module: str) -> list[Path]:
    root = ATP / module
    if not root.is_dir():
        return []
    return sorted(root.glob("*.yaml"))


def adb_bin() -> str:
    home = os.environ.get("ANDROID_HOME") or os.environ.get("ANDROID_SDK_ROOT")
    if home:
        cand = Path(home) / "platform-tools" / "adb.exe"
        if cand.exists():
            return str(cand)
    win = Path.home() / "AppData/Local/Android/Sdk/platform-tools/adb.exe"
    return str(win) if win.exists() else "adb"


def list_authorized_devices() -> list[str]:
    adb = adb_bin()
    try:
        p = subprocess.run([adb, "devices"], capture_output=True, text=True, timeout=30, check=False)
    except (OSError, subprocess.TimeoutExpired):
        return []
    out: list[str] = []
    for line in (p.stdout or "").splitlines():
        parts = line.strip().split()
        if len(parts) >= 2 and parts[1] == "device":
            out.append(parts[0])
    return out


def prepare(serial: str, *, parallel: bool = False) -> None:
    """Per-device hygiene. Avoid global ``adb forward --remove-all`` when parallel."""
    adb = adb_bin()
    cmds: list[list[str]] = []
    if not parallel:
        cmds.append(["forward", "--remove-all"])
    cmds.extend(
        [
            ["shell", "am", "force-stop", "dev.mobile.maestro"],
            ["shell", "am", "force-stop", "dev.mobile.maestro.test"],
            ["shell", "am", "force-stop", "com.android.settings"],
            ["shell", "input", "keyevent", "KEYCODE_HOME"],
        ]
    )
    for args in cmds:
        subprocess.run([adb, "-s", serial, *args], capture_output=True, timeout=30, check=False)


def run_flow(
    maestro: str,
    serial: str,
    flow: Path,
    timeout: int = 300,
    *,
    reinstall: bool = True,
) -> dict:
    t0 = time.time()
    # Global --device before test (docs.maestro.dev CLI); one serial per worker.
    # --no-ansi + file redirects avoid Windows Jansi UnsatisfiedLinkError (isatty on pipes).
    cmd = [maestro, "--no-ansi", "--device", serial, "test"]
    if reinstall:
        cmd.append("--reinstall-driver")
    cmd.append(str(flow))
    env = os.environ.copy()
    env["ANDROID_SERIAL"] = serial
    env["MAESTRO_CLI_NO_ANSI"] = "1"
    env["NO_COLOR"] = "1"
    env["JANSI_MODE"] = "force"
    env["TERM"] = "dumb"
    # Disable Jansi native isatty (common Windows crash when stdout is not a console).
    jansi_opts = "-Dorg.fusesource.jansi.Ansi.disable=true -Djansi.passthrough=true"
    prev = env.get("JAVA_TOOL_OPTIONS", "").strip()
    env["JAVA_TOOL_OPTIONS"] = f"{prev} {jansi_opts}".strip() if prev else jansi_opts
    out_dir = REPO / "reports" / "module_runs" / "_maestro_out"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{serial}_{os.getpid()}_{int(t0)}.log"
    try:
        with out_path.open("w", encoding="utf-8", errors="replace") as out_f:
            p = subprocess.run(
                cmd,
                cwd=str(REPO),
                stdout=out_f,
                stderr=subprocess.STDOUT,
                timeout=timeout,
                check=False,
                env=env,
            )
        out = out_path.read_text(encoding="utf-8", errors="replace") if out_path.exists() else ""
        try:
            out_path.unlink(missing_ok=True)
        except OSError:
            pass
        ok = p.returncode == 0
        reason = "" if ok else _fail_reason(out)
        return {
            "flow": flow.name,
            "path": str(flow.relative_to(REPO)),
            "status": "PASS" if ok else "FAIL",
            "exit_code": p.returncode,
            "failure_reason": reason,
            "execution_time_sec": round(time.time() - t0, 2),
            "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "device": serial,
        }
    except subprocess.TimeoutExpired:
        try:
            out_path.unlink(missing_ok=True)
        except OSError:
            pass
        return {
            "flow": flow.name,
            "path": str(flow.relative_to(REPO)),
            "status": "FAIL",
            "exit_code": -1,
            "failure_reason": f"timeout after {timeout}s",
            "execution_time_sec": round(time.time() - t0, 2),
            "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "device": serial,
        }
    except Exception as exc:  # noqa: BLE001
        try:
            out_path.unlink(missing_ok=True)
        except OSError:
            pass
        return {
            "flow": flow.name,
            "path": str(flow.relative_to(REPO)),
            "status": "FAIL",
            "exit_code": -2,
            "failure_reason": str(exc),
            "execution_time_sec": round(time.time() - t0, 2),
            "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "device": serial,
        }


def _fail_reason(out: str) -> str:
    lines = [ln.strip() for ln in (out or "").splitlines() if ln.strip()]
    for ln in reversed(lines[-40:]):
        low = ln.lower()
        if "assertion" in low or "not found" in low or "failed" in low or "error" in low:
            return ln[:300]
    return (lines[-1] if lines else "maestro non-zero exit")[:300]


def write_excel(module: str, summary: dict) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    xlsx = OUT / f"{module}_execution_report.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Execution"
    ws["A1"] = f"HP Sprocket Android - Module {module}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Flow", "Device", "Status", "Failure Reason", "Execution Time", "Timestamp"])
    pf = PatternFill("solid", fgColor="C6EFCE")
    ff = PatternFill("solid", fgColor="FFC7CE")
    for r in summary["rows"]:
        ws.append(
            [
                r["flow"],
                r.get("device", ""),
                r["status"],
                r.get("failure_reason", ""),
                r["execution_time_sec"],
                r["timestamp"],
            ]
        )
        cell = ws.cell(ws.max_row, 3)
        cell.fill = pf if r["status"] == "PASS" else ff
    sm = wb.create_sheet("Summary")
    for k in ("module", "total", "passed", "failed", "pass_percent", "execution_time_sec", "devices"):
        val = summary.get(k)
        if isinstance(val, list):
            val = ", ".join(str(x) for x in val)
        sm.append([k, val])
    wb.save(xlsx)
    return xlsx


def _parse_devices(raw: list[str]) -> list[str]:
    devices: list[str] = []
    for item in raw:
        for part in item.replace(";", ",").split(","):
            s = part.strip()
            if s and s not in devices:
                devices.append(s)
    return devices


def _run_one_with_retry(
    maestro: str,
    serial: str,
    flow: Path,
    timeout: int,
    *,
    reinstall_first: bool,
    parallel: bool,
) -> dict:
    prepare(serial, parallel=parallel)
    rec = run_flow(maestro, serial, flow, timeout=timeout, reinstall=reinstall_first)
    if rec["status"] == "PASS":
        return rec
    prepare(serial, parallel=parallel)
    need_reinstall = any(
        x in (rec.get("failure_reason") or "").lower()
        for x in (
            "install failed",
            "driver",
            "7001",
            "connection refused",
            "not connected",
            "unsatisfiedlinkerror",
            "jansi",
        )
    )
    rec2 = run_flow(maestro, serial, flow, timeout=timeout, reinstall=need_reinstall)
    if rec2["status"] == "PASS":
        rec2["failure_reason"] = f"passed on retry (first: {rec.get('failure_reason', '')[:120]})"
    return rec2


def run_parallel(
    maestro: str,
    devices: list[str],
    flows: list[Path],
    timeout: int,
) -> list[dict]:
    """Dynamic queue: each device worker pulls the next flow until empty."""
    work: queue.Queue[tuple[int, Path] | None] = queue.Queue()
    for i, flow in enumerate(flows):
        work.put((i, flow))
    for _ in devices:
        work.put(None)  # sentinel per worker

    results: dict[int, dict] = {}
    lock = threading.Lock()
    print_lock = threading.Lock()
    first_done = {s: False for s in devices}

    def worker(serial: str) -> None:
        while True:
            item = work.get()
            if item is None:
                work.task_done()
                break
            idx, flow = item
            with print_lock:
                print(f"[{idx + 1}/{len(flows)}] {flow.name} @ {serial} ...", flush=True)
            reinstall = not first_done[serial]
            first_done[serial] = True
            rec = _run_one_with_retry(
                maestro,
                serial,
                flow,
                timeout,
                reinstall_first=reinstall,
                parallel=True,
            )
            with lock:
                results[idx] = rec
            with print_lock:
                print(
                    f"  [{serial}] {rec['status']} ({rec['execution_time_sec']}s) "
                    f"{rec.get('failure_reason', '')[:100]}",
                    flush=True,
                )
            work.task_done()

    with ThreadPoolExecutor(max_workers=len(devices)) as pool:
        futs = [pool.submit(worker, serial) for serial in devices]
        for f in as_completed(futs):
            f.result()

    return [results[i] for i in range(len(flows)) if i in results]


def run_sequential(maestro: str, serial: str, flows: list[Path], timeout: int) -> list[dict]:
    rows: list[dict] = []
    prepare(serial, parallel=False)
    for i, flow in enumerate(flows, 1):
        print(f"[{i}/{len(flows)}] {flow.name} ...", flush=True)
        rec = _run_one_with_retry(
            maestro,
            serial,
            flow,
            timeout,
            reinstall_first=(i == 1),
            parallel=False,
        )
        rows.append(rec)
        print(f"  {rec['status']} ({rec['execution_time_sec']}s) {rec.get('failure_reason', '')[:120]}")
    return rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--module", required=True, help="ATP folder name e.g. home, photo-id")
    ap.add_argument(
        "--device",
        action="append",
        required=True,
        help="Device serial (repeat flag or comma-separate for parallel)",
    )
    ap.add_argument("--maestro", default=r"C:\Users\HP\maestro\maestro\bin\maestro.bat")
    ap.add_argument("--limit", type=int, default=0, help="Max flows (0=all)")
    ap.add_argument("--timeout", type=int, default=240)
    ap.add_argument(
        "--only-failed",
        action="store_true",
        help="Re-run only FAIL rows from reports/module_runs/<module>_summary.json and merge results",
    )
    ap.add_argument(
        "--from-flow",
        default="",
        help="Skip flows until this flow name (substring match) is reached",
    )
    ap.add_argument(
        "--parallel",
        action="store_true",
        help="Force parallel even with 1 device listed (no-op). Multi-device always parallel.",
    )
    args = ap.parse_args()

    devices = _parse_devices(args.device)
    if not devices:
        print("ERROR: no device serials")
        return 2
    online = set(list_authorized_devices())
    missing = [d for d in devices if d not in online]
    if missing:
        print(f"ERROR: device(s) not in adb 'device' state: {', '.join(missing)}")
        print(f"Online now: {', '.join(sorted(online)) or '(none)'}")
        return 2

    OUT.mkdir(parents=True, exist_ok=True)
    flows = discover_flows(args.module)
    prior_rows: list[dict] = []
    if args.only_failed:
        js_path = OUT / f"{args.module}_summary.json"
        if not js_path.is_file():
            print(f"ERROR: no summary to retry: {js_path}")
            return 2
        prior = json.loads(js_path.read_text(encoding="utf-8"))
        prior_rows = list(prior.get("rows") or [])
        fail_names = {r["flow"] for r in prior_rows if r.get("status") != "PASS"}
        flows = [f for f in flows if f.name in fail_names]
        print(f"[only-failed] retrying {len(flows)} failed flow(s) from {js_path.name}")
    if args.from_flow:
        needle = args.from_flow.strip().lower()
        idx = next((i for i, f in enumerate(flows) if needle in f.name.lower()), None)
        if idx is None:
            print(f"ERROR: --from-flow {args.from_flow!r} not found in module {args.module}")
            return 2
        flows = flows[idx:]
        print(f"[from-flow] starting at {flows[0].name} ({len(flows)} remaining)")
    if args.limit and args.limit > 0:
        flows = flows[: args.limit]
    if not flows:
        print(f"No flows in module: {args.module}")
        return 2

    mode = "parallel" if len(devices) > 1 else "sequential"
    print(
        f"===== MODULE {args.module} ({len(flows)} flows) devices={len(devices)} mode={mode} =====",
        flush=True,
    )
    print(f"Devices: {', '.join(devices)}", flush=True)

    t0 = time.time()
    if len(devices) > 1:
        rows = run_parallel(args.maestro, devices, flows, args.timeout)
    else:
        rows = run_sequential(args.maestro, devices[0], flows, args.timeout)

    if args.only_failed and prior_rows:
        by_name = {r["flow"]: r for r in prior_rows}
        for r in rows:
            by_name[r["flow"]] = r
        rows = [by_name[f.name] for f in discover_flows(args.module) if f.name in by_name]
        seen = {r["flow"] for r in rows}
        for r in prior_rows:
            if r["flow"] not in seen:
                rows.append(r)

    passed = sum(1 for r in rows if r["status"] == "PASS")
    failed = len(rows) - passed
    summary = {
        "module": args.module,
        "total": len(rows),
        "passed": passed,
        "failed": failed,
        "pass_percent": round((passed / len(rows)) * 100, 2) if rows else 0.0,
        "execution_time_sec": round(time.time() - t0, 2),
        "device": ",".join(devices),
        "devices": devices,
        "app": APP_ID,
        "rows": rows,
    }
    js = OUT / f"{args.module}_summary.json"
    js.write_text(json.dumps(summary, indent=2), encoding="utf-8")
    xlsx = write_excel(args.module, summary)

    index_path = OUT / "index.json"
    index = []
    if index_path.exists():
        try:
            index = json.loads(index_path.read_text(encoding="utf-8"))
        except Exception:
            index = []
    index = [x for x in index if x.get("module") != args.module]
    index.append({k: summary[k] for k in summary if k != "rows"})
    index_path.write_text(json.dumps(index, indent=2), encoding="utf-8")

    print("\n===== MODULE SUMMARY =====")
    print(json.dumps({k: summary[k] for k in summary if k != "rows"}, indent=2))
    print(f"report: {xlsx}")
    print(f"summary: {js}")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
